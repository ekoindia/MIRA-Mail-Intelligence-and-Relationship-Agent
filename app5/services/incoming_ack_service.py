"""
Acknowledgment-drafting workflow for periodic SBI-domain status pushes —
DRAFT ONLY. Built 2026-08-25 per explicit instruction, following the exact
same supervised-trial pattern as services/limit_forward_service.py.

Scope: the 2-year SBI-mail knowledge base review (1275 threads) found four
candidate categories with 0% historical reply rate. Reading the real subject
lines (not just the aggregate stats) before wiring this up dropped one of
them:
    SBI Data / Status Push               -- pure data-table push, kept
    Micro ATM Report                     -- standing instruction, kept
    BC-CSP Agreement & PVR Pendency Rpt  -- pure data-table push, kept
    Report Submission / Status           -- EXCLUDED, see below
"Report Submission / Status" looked safe from the aggregate stats (0%
replied) but its real subjects are a mixed bag that includes genuine,
sometimes urgent asks — "SUBMISSION OF CCPM INVOICE", "Required BC letter
for pre-arbitration", "ACTIVATION OF INACTIVE CSPS ... SUBSTITUTION PLAN —
URGENT" — plus replies landing on threads that are actually OUR OWN
automated report emails ("Re: Daily Progress Report | ..."). A generic
"received, noted" on any of those would read as acknowledging/actioning a
real request nobody has actually looked at. Excluded until it can be split
the same way Limit Approval was — by reading real bodies, not just subjects.

Everything else found in the review (Terminal/Device Issue, CSP Code
Allotment, Commission Query, Onboarding/KYC) has genuinely case-specific
replies and was never a candidate here.

Detection is triage_intent, not the fuzzy classify_reply_template keyword
score — triage_intent is already computed at ingestion and is deterministic
(ordered rules, first-match-wins), so it's the more reliable signal. The
matching IncomingReplyTemplate row (by category_name == triage_intent) only
supplies the actual reply CONTENT, so the wording stays editable from the
Suggestions/Templates UI without touching code.

HARD GUARANTEE: this module never sends mail. It only ever calls
gmail_service.create_outbound_draft, never send_message. A draft lands in
Gmail for a human to review and send — exactly like limit_forward_service.

REPLY, NOT A NEW EMAIL: added 2026-08-26 per explicit instruction — this
must land as a real reply on the original SBI thread (In-Reply-To/
References headers set from the original message's RFC Message-ID, plus
Gmail's own threadId), not a standalone "Re: ..." draft that only looks
like a reply by subject line. Reply to the original SENDER ONLY — never
reply-all, since these threads often carry many other recipients and only
SBI needs the acknowledgment. Every draft also CCs sbikiosk@eko.co.in,
matching the UNIVERSAL_CC every other automated mail in this app already
carries (see recipient_resolution_service.py).

READ THE DATA, DON'T BLANKET-REPLY: standing rule stated explicitly
2026-08-26 — "you always have to read data then decide which template to
use." SBI Data / Status Push already does this from the inline table.
Extended the same day to Micro ATM Report and BC-CSP Agreement & PVR
Pendency Report, whose data lives in an Excel attachment instead of the
body ("PFA...") — every such attachment carries a one-row-per-BC-partner
summary sheet (sheet NAME varies month to month — Sheet2, Sheet6, ... —
so it's found structurally, by a header row starting "BC NAME", not by
name) with Eko Bharat Ventures' own row among them. That row's real
numbers (inactive MATMs; expired agreements + pending PVRs) decide
"thank you, nothing pending" vs "noted, will action" — never a guess.
Autonomous end to end (no human reads the spreadsheet), but the HARD
GUARANTEE above is unchanged: still draft-only, still needs a human to
hit send.
"""
from __future__ import annotations

import html
import re

import openpyxl

from database.db import get_db
from database.incoming_models import IncomingAttachment, IncomingEmail, IncomingReplyTemplate
from services import gmail_service as gs
from services.incoming_service import _sender_email
from utils.logger import get_logger

logger = get_logger(__name__)

SBI_DOMAINS = ("sbi.co.in", "sbionline.onmicrosoft.com")

# Matches recipient_resolution_service.UNIVERSAL_CC — every automated mail
# in this app CCs the same address, ack replies included.
ACK_CC = "sbikiosk@eko.co.in"

ACK_INTENTS = [
    "SBI Data / Status Push",
    "Micro ATM Report",
    "BC-CSP Agreement & PVR Pendency Report",
    "Passbook Printer Report",
    "BC Commission Report",
]

# "SBI Data / Status Push" (Inactive Code Status) carries its data table
# INLINE in the email body — a row per BC partner, ours included. Micro
# ATM Report and BC-CSP Agreement & PVR Pendency Report instead say "PFA"
# and put the same shape of table in an Excel attachment (see
# _eko_row_from_workbook below) — all three now get read before a
# template is picked, none reply with a blanket "noted, will action".
EKO_ROW_MARKER = "eko bharat ventures"
SBI_STATUS_CLEAR_CATEGORY = "SBI Data / Status Push — Clear"
MICRO_ATM_CLEAR_CATEGORY = "Micro ATM Report — Clear"
PVR_PENDENCY_CLEAR_CATEGORY = "BC-CSP Agreement & PVR Pendency Report — Clear"
PASSBOOK_CLEAR_CATEGORY = "Passbook Printer Report — Clear"
COMMISSION_CLEAR_CATEGORY = "BC Commission Report — Clear"

# Sheets in a BC Commission Report attachment where a row naming Eko means
# a real problem (a failed/timed-out/un-attempted payment, or an inactive-
# CSP penalty) — checked in order, but ANY hit across ANY of them is
# enough to mean "needs action". Verified against a real 2026-08-10
# sample (Eko had rows in 3 of the 4).
COMMISSION_PROBLEM_SHEETS = [
    "Failed Payments", "Timed Out Payments", "Un-attempted Payments", "Inactive CSP Penalty",
]


def _is_sbi_sender(sender: str | None) -> bool:
    email = (_sender_email(sender) or "").lower()
    return any(email.endswith("@" + d) or email.endswith("." + d) for d in SBI_DOMAINS)


def _eko_inactive_code_count(body: str | None) -> int | None:
    """Reads Eko's own row out of an Inactive Code Status table and
    returns its Total Inactive Code count. Returns None (never guesses)
    if the row can't be found or doesn't parse — callers must treat None
    as "can't tell, use the default template", not as zero.

    The table survives HTML-to-text as one cell per line, in fixed column
    order (Total CSP Code, Inactive_31_90, Inactive_GT_90, Total Inactive
    Code, % Inactive) — including a blank line for an empty cell (SBI
    leaves 0-count cells blank rather than writing "0"), so the 5 lines
    right after the BC name line are always exactly this row's 5 values,
    positionally, verified against a real 2026-08-24 sample.
    """
    lines = (body or "").splitlines()
    idx = next((i for i, line in enumerate(lines) if EKO_ROW_MARKER in line.strip().lower()), None)
    if idx is None:
        return None
    values = lines[idx + 1: idx + 6]
    if len(values) < 4:
        return None
    total_inactive_raw = values[3].strip()
    if not total_inactive_raw:
        return 0
    try:
        return int(float(total_inactive_raw.replace(",", "")))
    except ValueError:
        return None


def _col_index(headers: list[str], *names: str) -> int | None:
    """Case/whitespace-tolerant column lookup — exact header text (spacing
    like "TOTAL INACTIVE >30D" vs a hypothetical ">30D TOTAL INACTIVE")
    isn't guaranteed identical month to month, so this matches by
    normalized substring against any of the given candidate names."""
    normalized = [(h or "").strip().upper() for h in headers]
    for name in names:
        target = name.strip().upper()
        for i, h in enumerate(normalized):
            if target in h:
                return i
    return None


def _eko_row_from_workbook(path: str) -> tuple[list[str], tuple] | tuple[None, None]:
    """Finds the one-row-per-BC-partner summary sheet in an SBI Excel
    attachment and returns (headers, Eko's own row values). The summary
    sheet's NAME varies between attachments (Sheet2, Sheet6, ...) — every
    workbook also carries much larger raw CSP-level detail sheets — so it
    is found structurally: the sheet whose first few rows contain a
    header row starting with "BC NAME". Returns (None, None) on any
    failure (bad file, sheet/row not found) — callers must treat that as
    "can't tell", never as a stand-in for zero.
    """
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not open Excel attachment %s: %s", path, exc)
        return None, None

    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
                if row and row[0] and str(row[0]).strip().upper() == "BC NAME":
                    header_row = [str(h).strip() if h is not None else "" for h in row]
                    break
            if header_row is None:
                continue
            in_data = False
            for row in ws.iter_rows(values_only=True):
                if not in_data:
                    if row and row[0] and str(row[0]).strip().upper() == "BC NAME":
                        in_data = True
                    continue
                if row and row[0] and EKO_ROW_MARKER in str(row[0]).strip().lower():
                    return header_row, row
            return header_row, None  # summary sheet found, but no Eko row in it
    finally:
        wb.close()
    return None, None


def _xlsx_attachment_path(db, incoming_email_id: int) -> str | None:
    att = (
        db.query(IncomingAttachment)
        .filter(IncomingAttachment.incoming_email_id == incoming_email_id)
        .filter(IncomingAttachment.stored_path.ilike("%.xlsx"))
        .first()
    )
    return att.stored_path if att else None


# A BC Commission Report attachment carries these two sheet names
# regardless of what the EMAIL SUBJECT says — real senders send this
# report under varying subject wording ("BC COMM REPORT FOR...", but
# also whatever else a given Chief Manager types that month), so the
# subject keyword rule alone misses real ones. The attachment's own
# structure is the reliable signal; verified against a real 2026-08-10
# sample (both sheets present).
COMMISSION_SIGNATURE_SHEETS = {"payment status", "circle wise break up"}


def _is_commission_report_attachment(path: str) -> bool:
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not open Excel attachment %s: %s", path, exc)
        return False
    try:
        sheet_names_lower = {s.strip().lower() for s in wb.sheetnames}
        return bool(sheet_names_lower & COMMISSION_SIGNATURE_SHEETS)
    finally:
        wb.close()


def reclassify_commission_by_attachment() -> int:
    """Finds SBI-sender incoming mail still unclassified ("other") whose
    xlsx attachment matches the BC Commission Report signature, and
    reclassifies it — content-based, not subject-based. Only ever touches
    rows with NO existing triage_intent (never overrides an existing,
    already-confident classification from a different rule). Safe to
    call repeatedly: already-classified rows are excluded from the query,
    so re-running only picks up genuinely new backlog. Called at the
    start of create_ack_drafts so this happens automatically, every run —
    no separate manual step."""
    updated = 0
    with get_db() as db:
        candidates = [
            (r.id, r.sender)
            for r in db.query(IncomingEmail).filter(
                IncomingEmail.recipient_kind == "to", IncomingEmail.triage_intent.is_(None)
            ).all()
        ]
    for eid, sender in candidates:
        if not _is_sbi_sender(sender):
            continue
        with get_db() as db:
            path = _xlsx_attachment_path(db, eid)
        if not path or not _is_commission_report_attachment(path):
            continue
        with get_db() as db:
            row = db.query(IncomingEmail).get(eid)
            if row and row.triage_intent is None:
                row.triage_tier = "info"
                row.triage_intent = "BC Commission Report"
                updated += 1
    if updated:
        logger.info("Reclassified %d incoming email(s) as BC Commission Report by attachment signature.", updated)
    return updated


def _micro_atm_needs_action(headers: list[str], values: tuple | None) -> bool | None:
    if values is None:
        return None
    idx = _col_index(headers, "TOTAL INACTIVE >30D", "TOTAL INACTIVE")
    if idx is None or idx >= len(values):
        return None
    try:
        return float(values[idx] or 0) > 0
    except (TypeError, ValueError):
        return None


def _pvr_pendency_needs_action(headers: list[str], values: tuple | None) -> bool | None:
    if values is None:
        return None
    idx_expired = _col_index(headers, "AGREEMENTS EXPIRED")
    idx_pending = _col_index(headers, "PVR PENDING")
    if idx_expired is None or idx_pending is None:
        return None
    if idx_expired >= len(values) or idx_pending >= len(values):
        return None
    try:
        expired = float(values[idx_expired] or 0)
        pending = float(values[idx_pending] or 0)
    except (TypeError, ValueError):
        return None
    return expired > 0 or pending > 0


def _excel_needs_action(db, incoming_email_id: int, decision_fn) -> bool | None:
    """Runs one of the decision functions above against the row's own
    Excel attachment. None propagates through every stage (no attachment,
    unreadable file, summary sheet not found, Eko's row not found, column
    not found) — always "can't tell", never a guessed answer."""
    path = _xlsx_attachment_path(db, incoming_email_id)
    if not path:
        return None
    headers, values = _eko_row_from_workbook(path)
    if headers is None:
        return None
    return decision_fn(headers, values)


def _passbook_needs_action(path: str) -> bool | None:
    """Passbook Printer Report attachments carry a "SUMM" sheet with a
    dual-block layout — the same NBC-partner table repeated twice
    side-by-side (an older snapshot on the left, the current one on the
    right, each dated in its own title row). Found structurally: the
    LAST "NBC NAME" column in the header row starts the current block,
    and Eko's row within that block is compared — enabled % below its
    own target % means real gap to close. Verified against a real
    2026-08-25 sample (Eko: 19.3% enabled vs a 38.95% target)."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not open Excel attachment %s: %s", path, exc)
        return None
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            header_row = None
            for row in ws.iter_rows(min_row=1, max_row=4, values_only=True):
                if row and any(c and str(c).strip().upper() == "NBC NAME" for c in row):
                    header_row = row
                    break
            if header_row is None:
                continue
            name_positions = [i for i, c in enumerate(header_row) if c and str(c).strip().upper() == "NBC NAME"]
            if not name_positions:
                continue
            start = name_positions[-1]
            sub_headers = [str(h).strip().upper() if h else "" for h in header_row[start:]]
            enable_idx = next((i for i, h in enumerate(sub_headers) if "ENABLE" in h), None)
            target_idx = next((i for i, h in enumerate(sub_headers) if h.startswith("TARGET")), None)
            if enable_idx is None or target_idx is None:
                return None
            for row in ws.iter_rows(values_only=True):
                name_val = row[start] if start < len(row) else None
                if name_val and EKO_ROW_MARKER in str(name_val).strip().lower():
                    try:
                        enable_pct = float(row[start + enable_idx] or 0)
                        target_pct = float(row[start + target_idx] or 0)
                    except (TypeError, ValueError, IndexError):
                        return None
                    return enable_pct < target_pct
            return None
    finally:
        wb.close()
    return None


def _commission_needs_action(path: str) -> bool | None:
    """A BC Commission Report attachment is a multi-sheet workbook, not a
    single summary table — checks whether Eko has any row on the
    "problem" sheets (failed/timed-out/un-attempted payments, inactive-
    CSP penalties). Any hit on any sheet means real issues to review;
    zero hits across every sheet the workbook actually has means clear.
    None only if the workbook has none of the expected sheets at all."""
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        logger.warning("Could not open Excel attachment %s: %s", path, exc)
        return None
    try:
        found_any_sheet = False
        for sheet_name in COMMISSION_PROBLEM_SHEETS:
            if sheet_name not in wb.sheetnames:
                continue
            found_any_sheet = True
            ws = wb[sheet_name]
            header = None
            for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
                header = [str(h).strip().upper() if h else "" for h in row]
            if not header:
                continue
            name_idx = next((i for i, h in enumerate(header) if "BC NAME" in h), None)
            if name_idx is None:
                continue
            for row in ws.iter_rows(min_row=2, values_only=True):
                val = row[name_idx] if name_idx < len(row) else None
                if val and "eko" in str(val).strip().lower():
                    return True
        return None if not found_any_sheet else False
    finally:
        wb.close()


def _render_subject(template: str, subject: str | None) -> str:
    subject = subject or "(no subject)"
    rendered = template.replace("{{Subject}}", subject)
    if not rendered.lower().startswith("re:"):
        rendered = f"Re: {rendered}"
    return rendered


def create_ack_drafts(max_items: int = 25, since=None) -> dict:
    """Create Gmail DRAFT acknowledgments for incoming SBI-domain mail in
    the ACK_INTENTS categories that don't have one yet. Never sends.
    Idempotent: ack_draft_id is set once a draft exists, so re-running
    skips those rows. Skips anything where a human has already replied on
    the thread (IncomingEmail.replied) — never draft a redundant ack on
    top of a real response.

    `since` bounds this to mail that ARRIVED after a given moment, same
    backlog-sweep guard as limit_forward_service.create_forward_sends.
    """
    reclassify_commission_by_attachment()

    summary = {"candidates": 0, "drafted": 0, "skipped_not_sbi": 0, "skipped_already_replied": 0, "errors": 0}

    with get_db() as db:
        templates = {
            t.category_name: t
            for t in db.query(IncomingReplyTemplate)
            .filter(
                IncomingReplyTemplate.category_name.in_(
                    ACK_INTENTS + [
                        SBI_STATUS_CLEAR_CATEGORY, MICRO_ATM_CLEAR_CATEGORY, PVR_PENDENCY_CLEAR_CATEGORY,
                        PASSBOOK_CLEAR_CATEGORY, COMMISSION_CLEAR_CATEGORY,
                    ]
                ),
                IncomingReplyTemplate.is_active.is_(True),
            )
            .all()
        }

        q = (
            db.query(IncomingEmail)
            .filter(
                IncomingEmail.recipient_kind == "to",
                IncomingEmail.triage_intent.in_(list(templates.keys())),
                IncomingEmail.ack_draft_id.is_(None),
            )
        )
        if since is not None:
            q = q.filter(IncomingEmail.received_at >= since)
        rows = (
            q.order_by(IncomingEmail.received_at.desc())
            .limit(max_items * 3)  # over-fetch: some filtered by guards below
            .all()
        )
        candidates = [
            (r.id, r.sender, r.subject, r.triage_intent, r.replied)
            for r in rows
        ]

    safe = []
    for cid, sender, subject, intent, replied in candidates:
        if not _is_sbi_sender(sender):
            summary["skipped_not_sbi"] += 1
            continue
        if replied:
            summary["skipped_already_replied"] += 1
            continue
        safe.append((cid, sender, subject, intent))
    safe = safe[:max_items]
    summary["candidates"] = len(safe)

    if not safe:
        logger.info("Incoming ack drafts: nothing to do. %s", summary)
        return summary

    try:
        service = gs.get_gmail_client()
    except Exception as exc:  # noqa: BLE001
        logger.error("Gmail client unavailable for ack drafts: %s", exc)
        summary["error"] = str(exc)
        return summary

    # Fetched once per batch (real network call) rather than per email — the
    # connected account's own configured Gmail signature, same as every
    # automated report email (services/email_service.py). Replaces a
    # hardcoded "Regards, ..." line with the real signature block.
    signature = gs.get_default_signature(service)

    for cid, sender, subject, intent in safe:
        try:
            with get_db() as db:
                row = db.query(IncomingEmail).get(cid)
                if row is None or row.ack_draft_id or row.replied:
                    continue
                template_category = intent
                if intent == "SBI Data / Status Push":
                    inactive_count = _eko_inactive_code_count(row.body_text)
                    # None (row/table not found) deliberately falls through
                    # to the default "will action" template — never assert
                    # "nothing pending" without actually having read a 0.
                    if inactive_count == 0:
                        template_category = SBI_STATUS_CLEAR_CATEGORY
                elif intent == "Micro ATM Report":
                    needs_action = _excel_needs_action(db, cid, _micro_atm_needs_action)
                    if needs_action is False:
                        template_category = MICRO_ATM_CLEAR_CATEGORY
                elif intent == "BC-CSP Agreement & PVR Pendency Report":
                    needs_action = _excel_needs_action(db, cid, _pvr_pendency_needs_action)
                    if needs_action is False:
                        template_category = PVR_PENDENCY_CLEAR_CATEGORY
                elif intent == "Passbook Printer Report":
                    path = _xlsx_attachment_path(db, cid)
                    needs_action = _passbook_needs_action(path) if path else None
                    if needs_action is False:
                        template_category = PASSBOOK_CLEAR_CATEGORY
                # BC Commission Report has no clear/no-action branch — per
                # explicit correction 2026-08-26, every commission report
                # genuinely needs review, so it always gets the
                # action-required template (template_category stays as the
                # plain intent, set above). _commission_needs_action and
                # COMMISSION_CLEAR_CATEGORY are kept unused in case
                # conditional detection comes back later.
                tpl = db.query(IncomingReplyTemplate).filter(
                    IncomingReplyTemplate.category_name == template_category, IncomingReplyTemplate.is_active.is_(True)
                ).first()
                if not tpl:
                    continue
                to_email = _sender_email(row.sender)
                if not to_email:
                    continue
                body_html = tpl.body_template + (signature or "")
                # Real reply threading, not a lookalike new email: the
                # original RFC Message-ID becomes In-Reply-To/References,
                # and Gmail's own thread id groups it in the same
                # conversation. to_email stays sender-only — no reply-all.
                original_message_id_header = None
                if row.gmail_message_id:
                    try:
                        original_message_id_header = gs.fetch_message_id_header(service, row.gmail_message_id)
                    except Exception as exc:  # noqa: BLE001
                        logger.warning("Could not fetch Message-ID header for incoming %s: %s", cid, exc)
                draft_id = gs.create_outbound_draft(
                    service,
                    to_email=to_email,
                    subject=_render_subject(tpl.subject_template, row.subject),
                    body_html=body_html,
                    attachment_path=None,
                    cc_emails=ACK_CC,
                    in_reply_to=original_message_id_header,
                    references=original_message_id_header,
                    thread_id=row.gmail_thread_id,
                )
                row.ack_draft_id = draft_id
                summary["drafted"] += 1
        except Exception as exc:  # noqa: BLE001
            logger.error("Ack draft failed for incoming %s: %s", cid, exc)
            summary["errors"] += 1

    logger.info("Incoming ack drafts: %s", summary)
    return summary
